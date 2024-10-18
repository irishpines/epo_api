from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from helpers.register_parser_functions import Patent


def date_format(date: datetime) -> str:
    return date.strftime("%d/%m/%Y")


def get_excel_from_template() -> Workbook:
    """
    Loads the template, which is an excel file with 5 sheets:
    "Instructions", "CASE_DATA", "NAME_DATA", "DESIGNATED_STATES", "CLASS_DATA"

    Returns the workbook object leaving template unchanged
    """
    wb = load_workbook(
        Path(
            r"C:\Users\remoteuser\Google Drive\Pythonscripts\epo_api\documentation\Standard_template Excel Conversion.xlsx"
        )
    )
    return wb


def initialize_wb(workbook: Workbook) -> Workbook:
    """
    deleting empty rows
    """
    wb = workbook
    for sheet in wb.worksheets:
        if sheet.title == "Instructions":
            continue
        sheet.delete_rows(2, sheet.max_row)
    return wb


def populate_excel_one_case(wb: Workbook, patent: Patent) -> Workbook:
    """
    Populates the CASE_DATA sheet with the data from the patent object
    """
    sheet = wb["CASE_DATA"]
    row = sheet.max_row + 1
    sheet[f"A{row}"].value = patent.ref  # Old_Case_No
    sheet[f"B{row}"].value = "Patent"  # Case_Type
    sheet[f"C{row}"].value = "EP"  # Country
    if patent.wo_application_number:
        if len(patent.priority) > 0:
            application_type = "PCT-Based With Priority"
        else:
            application_type = "PCT-Based Without Priority"
    else:
        if len(patent.priority) > 0:
            application_type = "With Priority"
        else:
            application_type = "Without Priority"
    sheet[f"D{row}"].value = application_type  # Application_Type
    sheet[f"E{row}"].value = "Seek Renewal Instructions"  # Service_Level
    sheet[f"F{row}"].value = patent.title  # Catchword
    # sheet[f"G{row}"].value ... Case_Number_Extension
    sheet[f"H{row}"].value = date_format(patent.filing_date)  # Application_Date
    sheet[f"I{row}"].value = patent.ep_application_number  # Application_No
    if len(patent.priority) == 1:
        sheet[f"J{row}"].value = date_format(patent.priority[0].date)  # Priority_Date
        sheet[f"K{row}"].value = patent.priority[0].number  # Priority_No
        sheet[f"L{row}"].value = patent.priority[0].country  # Priority_Country
    if len(patent.priority) > 1:
        sorted_priorities = sorted(patent.priority(key=lambda x: x.date))
        sheet[f"J{row}"].value = date_format(sorted_priorities[0].date)  # Priority_Date
        sheet[f"K{row}"].value = sorted_priorities[0].number  # Priority_No
        sheet[f"L{row}"].value = sorted_priorities[0].country  # Priority_Country
        add_priority_details = "Additional priorities:"
        for priority in sorted_priorities[1:]:
            add_priority_details += (
                f"{date_format(priority.date)} {priority.number} {priority.country} |"
            )
        sheet[f"AA{row}"].value = add_priority_details  # Notes_comments1
    sheet[f"M{row}"].value = date_format(patent.publication_date)  # Publication_Date
    sheet[f"N{row}"].value = patent.publication_number  # Publication_No
    # sheet[f"O{row}"].value ... Req_For_Examination
    # sheet[f"P{row}"].value ... Date_of_Use
    if patent.is_granted:
        sheet[f"Q{row}"].value = date_format(
            patent.grant_date
        )  # Registration_Grant_Date
        sheet[f"R{row}"].value = patent.publication_number  # Registration_No
    # sheet[f"S{row}"].value ... Next_renewal_annuity_Date
    # sheet[f"T{row}"].value ... This column hidden in version provided by Anthony: Official_Action_Due
    sheet[f"U{row}"].value = patent.applicants[0].unique_id  # Applicant
    # TODO: Ask how to handle multiple applicants in spreadsheet
    if len(patent.applicants) > 1:
        sheet[f"AB{row}"].value = "Co-Applicants: " + " | ".join(
            [applicant.unique_id for applicant in patent.applicants[1:]]
        )  # Notes_comments2
    # sheet[f"V{row}"].value ... Foreign_Agent
    sheet[
        f"W{row}"
    ].value = "17500"  # Kilpatrick Townsend & Stockton # Correspondence_Address
    # TODO: Decide how to populate correspondence address in future - maybe in spreadsheet??
    # TODO: Or prompt user which is messy?? Or leave blank for auto generation and populate later??
    sheet[f"X{row}"].value = "995579"  # Magic Leap # Account_Address
    # TODO: Same issue as for correspondence address
    sheet[f"Y{row}"].value = patent.inventors[0].unique_id  # Inventor_1
    if len(patent.inventors) > 1:
        sheet[f"Z{row}"].value = patent.inventors[1].unique_id  # Inventor_2
    if len(patent.inventors) > 2:
        co_inventor_str = "Co-Inventors: " + " | ".join(
            [inventor.unique_id for inventor in patent.inventors[2:]]
        )
        if sheet[f"AB{row}"].value:  # Already co-applicant info in there
            sheet[f"AB{row}"].value += " || " + co_inventor_str
        else:
            sheet[f"AB{row}"].value = co_inventor_str
    # sheet[f"AC{row}"].value ... This column hidden in version provided by Anthony: Next_Action_Date
    # sheet[f"AD{row}"].value ... This column hidden in version provided by Anthony: Next_Action
    # sheet[f"AE{row}"].value ... Trademark_appearance
    # sheet[f"AF{row}"].value ... Title_Description
    # sheet[f"AG{row}"].value ... TM_Type
    sheet[f"AH{row}"].value = "David Keane"  # Case_Responsible
    if patent.is_granted:
        sheet[f"AI{row}"].value = "Granted/Registered"  # Case_Status
    else:
        sheet[f"AI{row}"].value = "Pending"  # Case_Status
    sheet[f"AJ{row}"].value = patent.ref  # Agents_Ref
    wb.save(
        Path(
            rf"C:\Users\remoteuser\Google Drive\Pythonscripts\epo_api\output_files\Patricia_import_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        )
    )
